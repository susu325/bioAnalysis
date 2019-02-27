#%%
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
import openpyxl as oxl
import os
import seaborn as sns
# To improve printing of frame in notebook
from IPython.display import display, HTML

#%%
DEBUG = False

#%%
print(os.chdir('..\BioAnalysis'))
if DEBUG:
    print(os.getcwd())

# Idea is to run file inside your directory
# directory = ''
filename = "Van1Bio_analysis.xlsx"
sns.set_style('whitegrid')

#%%
wb = oxl.load_workbook(filename=filename, read_only=True)

#%%
# Get the Val1Bio16A part (range to be defined)
def ReadRangeToList(start, end):
    data_rows = []
    for row in ws[start + ':' + end]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    return data_rows

#%%
ws = wb['Van1Bio16']
column_names = ReadRangeToList('A2', 'F2')[0]  # Improve flattening


#%%
# Set the ranges of where the data is in each worksheet
start_list = ['AE3', 'J3', 'A2', 'AB3', 'A3']
end_list = ['AJ87', 'O91', 'F72', 'AG73', 'F131']
dfA = {}
ix_list = 0

for w_sheet in wb.sheetnames:
    data_a = []
    ws = wb[w_sheet]
    data_a = ReadRangeToList(start_list[ix_list], end_list[ix_list])
    df = pd.DataFrame(data_a, columns=column_names)
    ix_list = ix_list + 1
    dfA[w_sheet] = df


#%%
# A = pd.concat(dfA.values(), ignore_index=True)
A = pd.concat(dfA, ignore_index=False)
C = A.swaplevel()
C = C.reset_index(1)
repLevel = C['level_1'].str.replace('VanBio0a','00A').replace('Van1Bio8', '08A').replace('Van1Bio12', '12A').replace('Van1Bio16', '16A').replace('Van1Bio24','24A')
C['level_1'] = repLevel
C = C.sort_values(by=['level_1'])


#%%
D = C[C['gene'].str.startswith('SP_0284')]

#%%
D.to_csv('testo.csv')

#%%
image_dir = 'images/'
consens_dir = 'consensus/'

#%%
# grdf = C.groupby(['gene', 'mutation', 'annotation'])
grdf = C.groupby('gene')

#%%
DEBUG = True
#%%
if DEBUG:
    print(grdf.size())

#%%
for name, group in grdf:
    r = str(np.random.randint(100))
    print(name.split("\xa0")[0])
    # fig_name = os.path.join(image_dir, name.split("\xa0")[0] + '_' + r + + '.png')
    data_name = os.path.join(consens_dir, name.split("\xa0")[0] + '_' + r + '.csv')
    try:
        group[['level_1', 'timestap', 'mutation', 'freq', 'annotation', 'gene']].to_csv(data_name, index=False)
        # figure_gene = plot.figure()
        # figure_gene =  sns.catplot(x='level_1', y='freq',hue='gene', data=group, kind='bar')
        # plt.savefig(fig_name)

        # sns_plot = sns.catplot(x='level_1', y='freq',hue='gene', data=group, kind='bar')
        # sns_plot.savefig(fig_name)
        # group[['level_1', 'freq']].plot(kind='bar')
    except Exception as ex:
        print(ex)
        # pass
    # print(display(group))

#%%
print('---Finish---')



#-------------------------------------------------------------------------------#