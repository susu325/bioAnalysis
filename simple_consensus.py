"""
Simple consensus analysis - Create a consensus between samples in different
timepoints.

Usage:
    simple_consensus <input_file> ... [options]

"""
import argparse
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("TkAgg")
from matplotlib import pyplot as plt
import openpyxl as oxl
import os
import seaborn as sns


# Constants
verbose = False

# Globals
# start_ranges = ['AE3', 'J3', 'A2', 'AB3', 'A3']
# end_ranges = ['AJ87', 'O91', 'F72', 'AG73', 'F131']


def get_parser():
    parser = argparse.ArgumentParser(prog='simple consensus analysis',
            description='Creates a consensus between gene samples in different timepoints')
    parser.add_argument('input_file', type=str,
                        help='input file that contains the samples')
    parser.add_argument('output_directory', type=str, default='data',
                        help='output directory to store the results (default: data)')
    # parser.add_argument('-cn', '--column_names', type=str, nargs='+',
    #                     help=("names to be used as headers for the data. "
    #                     "Must match number of columns"))
    parser.add_argument('-r', '--ranges', type=str, nargs='+',
                        help='sets the starting ranges in each worksheet')
    parser.add_argument('-v', '--verbose', default=False, action='store_true',
                        help='activate verbose mode (default: off)')
    return parser


def read_range_to_list(ws, range_):
    """ Reads the values of every cell in range  """
    data_rows = []
    for row in ws[range_]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    return data_rows


def populate_group_data(wb, all_ranges, names):
    """ Reads each sheet in the specified range and adds it to a dict"""
    dfA = {}
    ix_list = 0
    for w_sheet in wb.sheetnames:
        data_a = []
        ws = wb[w_sheet]
        data_a = read_range_to_list(ws, all_ranges[ix_list])
        df = pd.DataFrame(data_a, columns=names)
        ix_list = ix_list + 1
        dfA[w_sheet] = df
    return dfA


def tidy_bio_data(dfA):
    A = pd.concat(dfA, ignore_index=False)
    C = A.swaplevel()
    C = C.reset_index(1)
    repLevel = C['level_1'].str.replace('VanBio0a','00A').replace('Van1Bio8', '08A').replace('Van1Bio12', '12A').replace('Van1Bio16', '16A').replace('Van1Bio24','24A')
    C['level_1'] = repLevel
    C = C.sort_values(by=['level_1'])
    return  C


def create_output_dirs(consensus_dir):
    if not os.path.exists(consensus_dir):
        os.mkdir(consensus_dir)

    os.chdir(consensus_dir)
    
    if not os.path.exists('data_dir'):
        os.mkdir('data_dir')

    if not os.path.exists('img_dir'):
        os.mkdir('img_dir')
    return 'data_dir', 'img_dir'


def generate_data_sets(C, consens_dir):
    grdf = C.groupby('gene')
    if verbose:
        print(grdf.size())

    data_dir, img_dir = create_output_dirs(consens_dir)
    
    for name, group in grdf:
        r = str(np.random.randint(100))
        if verbose:
            print(name.split("\xa0")[0])

        fig_name = os.path.join(img_dir, name.split("\xa0")[0] + '_' + r + '.png')
        data_name = os.path.join(data_dir, name.split("\xa0")[0] + '_' + r + '.csv')
    
        try:
            group[['level_1', 'timestamp', 'mutation', 'freq', 'annotation', 'gene']].to_csv(data_name, index=False)
            figure_gene = plt.figure()
            figure_gene =  sns.catplot(x='level_1', y='freq',hue='gene', data=group, kind='bar')
            plt.savefig(fig_name)
            plt.close()
            # sns_plot = sns.catplot(x='level_1', y='freq',hue='gene', data=group, kind='bar')
            # sns_plot.savefig(fig_name)
            # group[['level_1', 'freq']].plot(kind='bar')
        except Exception as ex:
            print(ex)
    return grdf, group, name


def validate_args():
    # validate_workbook()
    # validate_names()
    # validate_ranges()
    pass


def main():
    """ Read input arguments """
    parser = get_parser()
    args = parser.parse_args()
    # validate_args()
    verbose = args.verbose
    if verbose:
        print(args)
    filename = args.input_file
    column_names = ['timestamp', 'mutation', 'freq', 'annotation', 'gene', 'description']
    out_dir = args.output_directory
    # Add beautiful style
    sns.set_style('whitegrid')
    wb = oxl.load_workbook(filename=filename, read_only=True)
    ranges = args.ranges

    dfA = populate_group_data(wb, ranges, column_names)
    if verbose:
        for d in dfA:
            print(d)
            print(dfA[d].head())

    C = tidy_bio_data(dfA)
    grdf, group, name = generate_data_sets(C, out_dir)


if __name__ == '__main__':
    main()
    print('---Done---')


#D = C[C['gene'].str.startswith('SP_0284')]
